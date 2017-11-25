from django.shortcuts import render
from .models import Photo, Person, Vote, PersonForm, VoteForm
from survey import settings
from random import shuffle

import openpyxl

# Create your views here.


def photo_survey(request, univcode):
    gist = list(Photo.objects.filter(univ_code='G'))
    jeon = list(Photo.objects.filter(univ_code='J'))
    shuffle(gist)
    shuffle(jeon)
    photos = gist[:10]+jeon[:10]
    shuffle(photos)

    return render(request, 'index.html', 
        {'univcode': univcode,
        'studentcode': 0 if univcode == 'G' else 0,
        'photos': photos,
        'names': ' '.join([photo.filename for photo in photos])})

def submit(request):
    if request.method == "POST":
        post = request.POST #univ_code, studentcode, (photo)names, scores, answers

        if int(post['studentcode']) != 0:
            studentcode = int(post['studentcode'])
        else:
            studentcode = len(Person.objects.filter(univ_code=str(post['univ_code'])))
        names = post['names'].split()
        scores = list(map(int, post['scores'].split()))

        personcheck = Person.objects.filter(studentcode=studentcode)
        person = None

        if not personcheck:
            personform = PersonForm({'univ_code': post['univ_code'], 'studentcode': studentcode})
            if personform.is_valid():
                person = personform.save(commit=False)
                person.save()
        else:
            person = personcheck[0]

        scores.reverse() #stack -> queue
        for name in names:
            voteform = VoteForm({})
            if voteform.is_valid():
                vote = voteform.save(commit=False)
                vote.score1 = scores.pop()
                vote.score2 = scores.pop()
                vote.score3 = scores.pop()
                vote.voter = person
                vote.photo = Photo.objects.get(filename=name)
                vote.save()
            
        #answers = post['answers'].split()

    return render(request, 'redirect.html')

def create_sheet(request):
    wb = openpyxl.Workbook()

    #sheet 1 : score mean by photo
    ws = wb.active
    photos = Photo.objects.all()


    ws.append(['사진', '소속대학', '성별', '응답자 수', '단정성', '세련성', '활동성'])

    for photo in photos:
        ws.append([photo.filename, photo.univ_code, photo.gender, len(photo.GetVotes()), *photo.GetMean()])


    #sheet 2 : votes
    ws = wb.create_sheet(title="투표")
    people = Person.objects.all()
    votes = Vote.objects.all()

    ws.append(['학번', '소속대학', '성별', '사진', '단정성', '세련성', '활동성'])

    for person in people:
        info = [person.studentcode, person.univ_code, person.gender]

        for vote in votes.filter(voter=person):
            ws.append([*info, vote.photo.filename, vote.score1, vote.score2, vote.score3])
            info = ['','','']

    #sheet 3
    ws = wb.create_sheet(title="")

    

    ws.append(['사진 수', len(photos)])
    ws.append(['투표자 수', len(people)])

    wb.save('static/output/result.xlsx')

    return render(request, 'result.html')